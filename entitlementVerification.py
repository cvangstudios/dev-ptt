import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os
from datetime import datetime
from collections import defaultdict

def analyze_entitlements(input_file):
    """
    Analyzes entitlements from Excel file and creates comprehensive analysis.
    
    Args:
        input_file (str): Path to input Excel file with people as columns and entitlements as values
    """
    
    # Read the Excel file
    print(f"Reading {input_file}...")
    df = pd.read_excel(input_file, header=0)
    
    # Get people names from column headers
    people = [col for col in df.columns if pd.notna(col) and col.strip() != '']
    print(f"Found {len(people)} people: {people}")
    
    # Collect entitlements for each person
    person_entitlements = {}
    
    for person in people:
        # Get all non-null, non-empty values in this person's column
        entitlements = []
        if person in df.columns:
            person_data = df[person].dropna()
            person_data = person_data[person_data.astype(str).str.strip() != '']
            entitlements = person_data.astype(str).str.strip().tolist()
        
        person_entitlements[person] = list(set(entitlements))  # Remove duplicates
        print(f"{person}: {len(person_entitlements[person])} entitlements")
    
    # Get all unique entitlements across everyone
    all_entitlements = set()
    for entitlements in person_entitlements.values():
        all_entitlements.update(entitlements)
    
    print(f"Total unique entitlements across all people: {len(all_entitlements)}")
    
    # 1. Find COMMON entitlements (ones that ALL people have)
    common_entitlements = []
    for entitlement in all_entitlements:
        if all(entitlement in person_entitlements[person] for person in people):
            common_entitlements.append(entitlement)
    
    common_entitlements.sort()
    print(f"Common entitlements (all have): {len(common_entitlements)}")
    
    # 2. Find DIFFERENT entitlements per person (ones they have that not everyone has)
    different_entitlements = {}
    for person in people:
        different_entitlements[person] = [
            ent for ent in person_entitlements[person] 
            if ent not in common_entitlements
        ]
        different_entitlements[person].sort()
    
    # 3. Find PERSON-EXCLUSIVE entitlements (ones that only ONE person has)
    exclusive_entitlements = {person: [] for person in people}
    
    for entitlement in all_entitlements:
        people_with_entitlement = [
            person for person in people 
            if entitlement in person_entitlements[person]
        ]
        
        if len(people_with_entitlement) == 1:
            exclusive_person = people_with_entitlement[0]
            exclusive_entitlements[exclusive_person].append(entitlement)
    
    # Sort exclusive entitlements
    for person in people:
        exclusive_entitlements[person].sort()
    
    # Create the analysis workbook
    print("Creating analysis workbook...")
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, color="000000")
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    
    # Write summary data
    summary_data = [
        ["Analysis Summary", ""],
        ["Total People", len(people)],
        ["Total Unique Entitlements", len(all_entitlements)],
        ["Common Entitlements (all have)", len(common_entitlements)],
        ["", ""],
        ["Different Entitlements per Person:", ""],
    ]
    
    for person in people:
        summary_data.append([f"  {person}", len(different_entitlements[person])])
    
    summary_data.extend([
        ["", ""],
        ["Exclusive Entitlements per Person:", ""],
    ])
    
    for person in people:
        summary_data.append([f"  {person}", len(exclusive_entitlements[person])])
    
    # Write to sheet
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1 or "per Person:" in str(value):
                cell.font = subheader_font
                cell.fill = subheader_fill
    
    # Auto-adjust column width
    for col_letter in ['A', 'B']:
        max_length = 0
        for row in ws_summary.iter_rows(min_col=ord(col_letter)-64, max_col=ord(col_letter)-64):
            for cell in row:
                if hasattr(cell, 'value') and cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws_summary.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Sheet 2: Common Entitlements
    ws_common = wb.create_sheet("Common Entitlements")
    
    ws_common.cell(1, 1, "Common Entitlements (All People Have These)").font = header_font
    ws_common.cell(1, 1).fill = header_fill
    ws_common.merge_cells('A1:B1')
    
    ws_common.cell(2, 1, "Entitlement").font = subheader_font
    ws_common.cell(2, 1).fill = subheader_fill
    ws_common.cell(2, 2, "People Count").font = subheader_font
    ws_common.cell(2, 2).fill = subheader_fill
    
    for idx, entitlement in enumerate(common_entitlements, 3):
        ws_common.cell(idx, 1, entitlement)
        ws_common.cell(idx, 2, len(people))
    
    # Auto-adjust column width
    for col_letter in ['A', 'B']:
        max_length = 0
        for row in ws_common.iter_rows(min_col=ord(col_letter)-64, max_col=ord(col_letter)-64):
            for cell in row:
                if hasattr(cell, 'value') and cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws_common.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Sheet 3: Different Entitlements
    ws_different = wb.create_sheet("Different Entitlements")
    
    # Create header
    ws_different.cell(1, 1, "Different Entitlements by Person").font = header_font
    ws_different.cell(1, 1).fill = header_fill
    ws_different.merge_cells(f'A1:{chr(65 + len(people))}1')
    
    # Column headers
    for idx, person in enumerate(people, 2):
        cell = ws_different.cell(2, idx, person)
        cell.font = subheader_font
        cell.fill = subheader_fill
    
    ws_different.cell(2, 1, "Row").font = subheader_font
    ws_different.cell(2, 1).fill = subheader_fill
    
    # Find max entitlements for any person
    max_different = max(len(different_entitlements[person]) for person in people)
    
    # Write different entitlements
    for row_idx in range(max_different):
        ws_different.cell(row_idx + 3, 1, row_idx + 1)
        for col_idx, person in enumerate(people, 2):
            if row_idx < len(different_entitlements[person]):
                ws_different.cell(row_idx + 3, col_idx, different_entitlements[person][row_idx])
    
    # Auto-adjust column width
    max_col_letter = chr(65 + len(people))
    for i in range(len(people) + 1):
        col_letter = chr(65 + i)
        max_length = 0
        for row in ws_different.iter_rows(min_col=i+1, max_col=i+1):
            for cell in row:
                if hasattr(cell, 'value') and cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws_different.column_dimensions[col_letter].width = min(max_length + 2, 30)
    
    # Sheet 4: Exclusive Entitlements
    ws_exclusive = wb.create_sheet("Exclusive Entitlements")
    
    # Create header
    ws_exclusive.cell(1, 1, "Exclusive Entitlements (Only One Person Has)").font = header_font
    ws_exclusive.cell(1, 1).fill = header_fill
    ws_exclusive.merge_cells(f'A1:{chr(65 + len(people))}1')
    
    # Column headers
    for idx, person in enumerate(people, 2):
        cell = ws_exclusive.cell(2, idx, person)
        cell.font = subheader_font
        cell.fill = subheader_fill
    
    ws_exclusive.cell(2, 1, "Row").font = subheader_font
    ws_exclusive.cell(2, 1).fill = subheader_fill
    
    # Find max exclusive entitlements for any person
    max_exclusive = max(len(exclusive_entitlements[person]) for person in people) if any(exclusive_entitlements.values()) else 0
    
    if max_exclusive > 0:
        # Write exclusive entitlements
        for row_idx in range(max_exclusive):
            ws_exclusive.cell(row_idx + 3, 1, row_idx + 1)
            for col_idx, person in enumerate(people, 2):
                if row_idx < len(exclusive_entitlements[person]):
                    ws_exclusive.cell(row_idx + 3, col_idx, exclusive_entitlements[person][row_idx])
    else:
        ws_exclusive.cell(3, 1, "No exclusive entitlements found")
        ws_exclusive.merge_cells(f'A3:{chr(65 + len(people))}3')
    
    # Auto-adjust column width
    for i in range(len(people) + 1):
        col_letter = chr(65 + i)
        max_length = 0
        for row in ws_exclusive.iter_rows(min_col=i+1, max_col=i+1):
            for cell in row:
                if hasattr(cell, 'value') and cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws_exclusive.column_dimensions[col_letter].width = min(max_length + 2, 30)
    
    # Sheet 5: Complete Matrix
    ws_matrix = wb.create_sheet("Complete Matrix")
    
    # Create header
    ws_matrix.cell(1, 1, "Complete Entitlement Matrix").font = header_font
    ws_matrix.cell(1, 1).fill = header_fill
    ws_matrix.merge_cells(f'A1:{chr(66 + len(people))}1')
    
    # Column headers
    ws_matrix.cell(2, 1, "Entitlement").font = subheader_font
    ws_matrix.cell(2, 1).fill = subheader_fill
    
    for idx, person in enumerate(people, 2):
        cell = ws_matrix.cell(2, idx, person)
        cell.font = subheader_font
        cell.fill = subheader_fill
    
    ws_matrix.cell(2, len(people) + 2, "Total Count").font = subheader_font
    ws_matrix.cell(2, len(people) + 2).fill = subheader_fill
    
    # Write matrix data - sort by count (least to most people who have it)
    entitlement_counts = []
    for entitlement in all_entitlements:
        count = sum(1 for person in people if entitlement in person_entitlements[person])
        entitlement_counts.append((entitlement, count))
    
    # Sort by count (ascending), then alphabetically for ties
    entitlement_counts.sort(key=lambda x: (x[1], x[0]))
    
    for row_idx, (entitlement, _) in enumerate(entitlement_counts, 3):
        ws_matrix.cell(row_idx, 1, entitlement)
        
        count = 0
        for col_idx, person in enumerate(people, 2):
            has_entitlement = entitlement in person_entitlements[person]
            ws_matrix.cell(row_idx, col_idx, "✓" if has_entitlement else "")
            if has_entitlement:
                count += 1
        
        ws_matrix.cell(row_idx, len(people) + 2, count)
    
    # Auto-adjust column width
    for i in range(len(people) + 2):
        col_letter = chr(65 + i)
        max_length = 0
        for row in ws_matrix.iter_rows(min_col=i+1, max_col=i+1):
            for cell in row:
                if hasattr(cell, 'value') and cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws_matrix.column_dimensions[col_letter].width = min(max_length + 2, 30)
    
    # Save the workbook with backup handling
    output_file = "entitlement_analysis.xlsx"
    
    # Check if output file already exists and backup if needed
    if os.path.exists(output_file):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = f"entitlement_analysis.old.{timestamp}.xlsx"
        try:
            os.rename(output_file, backup_file)
            print(f"Previous analysis backed up as: {backup_file}")
        except Exception as e:
            print(f"Warning: Could not backup existing file: {e}")
            print("Proceeding to overwrite existing file...")
    
    wb.save(output_file)
    
    print(f"\nAnalysis complete! Results saved to: {output_file}")
    print(f"\nSummary:")
    print(f"- Total people: {len(people)}")
    print(f"- Total unique entitlements: {len(all_entitlements)}")
    print(f"- Common entitlements (all have): {len(common_entitlements)}")
    print(f"- Different entitlements per person: {[len(different_entitlements[p]) for p in people]}")
    print(f"- Exclusive entitlements per person: {[len(exclusive_entitlements[p]) for p in people]}")
    
    return {
        'people': people,
        'all_entitlements': sorted(list(all_entitlements)),
        'common_entitlements': common_entitlements,
        'different_entitlements': different_entitlements,
        'exclusive_entitlements': exclusive_entitlements
    }

if __name__ == "__main__":
    input_file = "entitlements.xlsx"
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print("="*60)
        print("ERROR: Input file not found!")
        print("="*60)
        print(f"The script is looking for: '{input_file}'")
        print(f"Current directory: {os.getcwd()}")
        print("\nPlease ensure the following:")
        print("1. The file 'entitlements.xlsx' exists in the same folder as this script")
        print("2. The file name is spelled correctly (case-sensitive)")
        print("3. The file is not open in Excel (close it if open)")
        print("\nExiting...")
        sys.exit(1)
    
    # Check file permissions
    if not os.access(input_file, os.R_OK):
        print("="*60)
        print("ERROR: Cannot read input file!")
        print("="*60)
        print(f"File '{input_file}' exists but cannot be read.")
        print("Please check file permissions and ensure it's not open in Excel.")
        print("\nExiting...")
        sys.exit(1)
    
    try:
        print("="*60)
        print("ENTITLEMENT ANALYSIS STARTING")
        print("="*60)
        print(f"Input file: {input_file}")
        print(f"Current time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()
        
        results = analyze_entitlements(input_file)
        
        print("="*60)
        print("ANALYSIS COMPLETED SUCCESSFULLY!")
        print("="*60)
        
    except FileNotFoundError:
        print("="*60)
        print("ERROR: File not found during processing!")
        print("="*60)
        print(f"Could not find '{input_file}' during analysis.")
        print("The file may have been moved or deleted while processing.")
        sys.exit(1)
        
    except PermissionError:
        print("="*60)
        print("ERROR: Permission denied!")
        print("="*60)
        print("Cannot access the file. Please check:")
        print("1. File is not open in Excel")
        print("2. You have read/write permissions")
        print("3. File is not locked by another process")
        sys.exit(1)
        
    except pd.errors.EmptyDataError:
        print("="*60)
        print("ERROR: Empty or invalid Excel file!")
        print("="*60)
        print(f"The file '{input_file}' appears to be empty or corrupted.")
        print("Please check that it contains valid data with:")
        print("1. People names in the first row")
        print("2. Entitlements in the rows below")
        sys.exit(1)
        
    except Exception as e:
        print("="*60)
        print("ERROR: Unexpected error during analysis!")
        print("="*60)
        print(f"Error details: {str(e)}")
        print("\nFull error traceback:")
        import traceback
        traceback.print_exc()
        print("\nIf this error persists, please check:")
        print("1. Excel file format and structure")
        print("2. Data contains valid entries")
        print("3. No special characters causing issues")
        sys.exit(1)






Option Explicit

' Module-level variable to store the previously highlighted row
Private PreviousRow As Long

' This subroutine runs whenever the selection changes in the worksheet
Private Sub Worksheet_SelectionChange(ByVal Target As Range)
    
    ' Exit if multiple cells are selected to avoid issues
    If Target.Cells.Count > 1 Then Exit Sub
    
    ' Clear highlighting from the previous row (if any)
    If PreviousRow > 0 Then
        With Rows(PreviousRow).Interior
            .ColorIndex = xlNone ' Remove background color
            .Pattern = xlNone
        End With
    End If
    
    ' Highlight the current row
    With Target.EntireRow.Interior
        .Color = RGB(255, 255, 200) ' Light yellow highlight
        .Pattern = xlSolid
    End With
    
    ' Store the current row number for next time
    PreviousRow = Target.Row
    
End Sub

' Optional: Add this to clear highlighting when leaving the worksheet
Private Sub Worksheet_Deactivate()
    
    ' Clear highlighting when switching to another worksheet
    If PreviousRow > 0 Then
        With Rows(PreviousRow).Interior
            .ColorIndex = xlNone
            .Pattern = xlNone
        End With
        PreviousRow = 0
    End If
    
End Sub
