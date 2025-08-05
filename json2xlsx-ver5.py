#!/usr/bin/env python3
"""
Generic API JSON Scraper with Excel Export
Version: 1.0
Description: Fetches JSON data from API endpoints using serial numbers,
             creates pretty-printed JSON files, and consolidates data into Excel.
             
Usage: python script.py
Requirements: 'serials.txt' file in same directory as script
             pip install openpyxl
"""

import subprocess
import json
import csv
import os
import sys
import getpass
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.hyperlink import Hyperlink
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Install with: pip install openpyxl")
    print("   Falling back to CSV format...")

# Script Configuration
SCRIPT_VERSION = "1.0"
BASE_URL_TEMPLATE = "https://acme.com/sn="  # Modify this for your API endpoint

def read_serial_numbers(file_path):
    """
    Read serial numbers from text file.
    First line is project name, remaining lines are serial numbers.
    
    Args:
        file_path (str): Path to text file containing serial numbers
        
    Returns:
        tuple: (project_name, list_of_serial_numbers)
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = [line.strip() for line in f.readlines() if line.strip()]
        
        if not lines:
            raise ValueError("File is empty")
            
        project_name = lines[0]
        serial_numbers = lines[1:]
        
        print(f"Project: {project_name}")
        print(f"Found {len(serial_numbers)} serial numbers")
        
        return project_name, serial_numbers
        
    except FileNotFoundError:
        print(f"❌ Error: File '{file_path}' not found")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        sys.exit(1)

def check_configuration_intent(serial_numbers, base_url):
    """
    Pre-check all serial numbers for valid ConfigurationIntent data.
    
    Args:
        serial_numbers (list): List of serial numbers to check
        base_url (str): Base URL template for API
        
    Returns:
        tuple: (valid_serials, invalid_serials)
    """
    print("🔍 Pre-checking all serials for valid ConfigurationIntent data...")
    print("="*60)
    
    valid_serials = []
    invalid_serials = []
    
    for i, serial_number in enumerate(serial_numbers, 1):
        print(f"[{i}/{len(serial_numbers)}] Checking: {serial_number}", end=" ")
        
        url = f"{base_url}{serial_number}"
        
        try:
            # Quick API call to check ConfigurationIntent
            result = subprocess.run([
                'curl',
                '-s',  # Silent mode
                '-L',  # Follow redirects
                '--max-time', '30',  # 30 second timeout
                '--fail',  # Fail on HTTP errors
                '--negotiate',  # Enable SPNEGO/Negotiate authentication (Kerberos/NTLM)
                '--user', ':',  # Use current user credentials (empty user:pass triggers system auth)
                '--ntlm',  # Enable NTLM authentication for Windows domains
                '--anyauth',  # Let curl pick the best auth method available
                '--insecure',  # Skip SSL certificate verification (for testing)
                '--ssl-no-revoke',  # Don't check SSL certificate revocation (Windows)
                '--tlsv1.2',  # Force TLS 1.2 or higher
                url
            ], capture_output=True, text=True, check=True, encoding='utf-8')
            
            # Parse JSON response
            data = json.loads(result.stdout)
            
            # Check ConfigurationIntent
            config_intent = data.get('ConfigurationIntent')
            if config_intent is None:
                print("❌ (ConfigurationIntent is null)")
                invalid_serials.append(serial_number)
            elif not isinstance(config_intent, dict):
                print(f"❌ (ConfigurationIntent is {type(config_intent).__name__}, not dict)")
                invalid_serials.append(serial_number)
            else:
                print("✅ (Valid)")
                valid_serials.append(serial_number)
                
        except subprocess.CalledProcessError as e:
            print(f"❌ (API Error: Status {e.returncode})")
            invalid_serials.append(serial_number)
        except json.JSONDecodeError as e:
            print("❌ (Invalid JSON)")
            invalid_serials.append(serial_number)
        except Exception as e:
            print(f"❌ (Error: {str(e)[:50]}...)")
            invalid_serials.append(serial_number)
    
    # Summary of pre-check
    print("="*60)
    print("📊 PRE-CHECK SUMMARY:")
    print(f"✅ Valid serials: {len(valid_serials)}")
    print(f"❌ Invalid serials: {len(invalid_serials)}")
    
    if invalid_serials:
        print(f"\n🚫 Serials with null/invalid ConfigurationIntent:")
        for serial in invalid_serials:
            print(f"   • {serial}")
        print(f"\n⚠️  These {len(invalid_serials)} serials will be skipped during processing.")
    
    if valid_serials:
        print(f"\n✅ Valid serials to process:")
        for serial in valid_serials:
            print(f"   • {serial}")
        print(f"\n🚀 Will process {len(valid_serials)} valid serials...")
    else:
        print(f"\n❌ No valid serials found! All serials have invalid ConfigurationIntent.")
        return valid_serials, invalid_serials
    
    print("="*60)
    input("Press Enter to continue with processing valid serials...")
    print()
    
    return valid_serials, invalid_serials
def fetch_json_data(serial_number, base_url):
    """
    Fetch JSON data from API using curl with system authentication.
    
    Args:
        serial_number (str): Serial number to query
        base_url (str): Base URL template for API
        
    Returns:
        dict or None: JSON data if successful, None if failed
    """
    url = f"{base_url}{serial_number}"
    
    try:
        print(f"Fetching data for SN: {serial_number}")
        
        # Execute curl command with system authentication
        curl_cmd = [
            'curl',
            '-s',  # Silent mode
            '-L',  # Follow redirects
            '--max-time', '30',  # 30 second timeout
            '--fail',  # Fail on HTTP errors
            '--negotiate',  # Enable SPNEGO/Negotiate authentication (Kerberos/NTLM)
            '--user', ':',  # Use current user credentials (empty user:pass triggers system auth)
            '--ntlm',  # Enable NTLM authentication for Windows domains
            '--anyauth',  # Let curl pick the best auth method available
            '--insecure',  # Skip SSL certificate verification (for testing)
            '--ssl-no-revoke',  # Don't check SSL certificate revocation (Windows)
            '--tlsv1.2',  # Force TLS 1.2 or higher
            url
        ]
        
        result = subprocess.run(curl_cmd, capture_output=True, text=True, check=True, encoding='utf-8')
        
        # Parse JSON response
        data = json.loads(result.stdout)
        return data
        
    except subprocess.CalledProcessError as e:
        print(f"❌ API request failed for {serial_number}: HTTP error (Status: {e.returncode})")
        print(f"   This might be an authentication issue. Check your domain credentials.")
        return None
    except json.JSONDecodeError as e:
        print(f"❌ Invalid JSON response for {serial_number}: {e}")
        return None
    except Exception as e:
        print(f"❌ Error fetching {serial_number}: {e}")
        return None

def create_pretty_json_file(serial_number, data, project_name, output_dir):
    """
    Create a pretty-printed JSON file with header information.
    
    Args:
        serial_number (str): Serial number
        data (dict): JSON data from API
        project_name (str): Project name for header
        output_dir (Path): Output directory path
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Create header information
    header_info = {
        "metadata": {
            "project_name": project_name,
            "serial_number": serial_number,
            "timestamp": timestamp,
            "script_version": SCRIPT_VERSION,
            "data_source": f"{BASE_URL_TEMPLATE}{serial_number}"
        },
        "api_data": data
    }
    
    # Save to JSON file
    file_path = output_dir / f"{serial_number}.json"
    
    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(header_info, f, indent=2, ensure_ascii=False)
        print(f"✅ Saved JSON: {file_path}")
        
    except Exception as e:
        print(f"❌ Error saving JSON for {serial_number}: {e}")

def extract_csv_data(serial_number, data, valid_serials):
    """
    Extract specific data fields for CSV export.
    Uses conditional logic: full extraction for valid ConfigurationIntent, basic extraction for null.
    
    Args:
        serial_number (str): Serial number
        data (dict): JSON data from API
        valid_serials (list): List of serials with valid ConfigurationIntent
        
    Returns:
        dict or None: Flattened data for CSV row, or None if extraction fails
    """
    
    try:
        # Check if data is None or empty
        if data is None:
            print(f"⚠️  No data received for {serial_number} - API returned None")
            return None
            
        if not isinstance(data, dict):
            print(f"⚠️  Invalid data type for {serial_number} - expected dict, got {type(data)}")
            return None
        
        csv_row = {
            'serial_number': serial_number
        }
        
        # =================================================================
        # CONDITIONAL EXTRACTION BASED ON CONFIGURATIONINTENT STATUS
        # =================================================================
        
        # Get ConfigurationIntent safely
        config_intent = data.get('ConfigurationIntent')
        
        if serial_number in valid_serials and config_intent is not None:
            # FULL EXTRACTION - Valid ConfigurationIntent
            csv_row['config_status'] = 'Valid'
            
            print(f"   Using FULL extraction for {serial_number} (valid ConfigurationIntent)")
            
            # CUSTOMIZE THIS SECTION - Add your specific field extractions for valid ConfigurationIntent devices
            # Safe extraction with null checking
            if isinstance(config_intent, dict):
                intent_data = config_intent.get('IntentData', {})
                
                # Add your specific field extractions here
                # Example:
                # csv_row.update({
                #     'your_field_1': intent_data.get('your_field_1', ''),
                #     'your_field_2': intent_data.get('your_field_2', ''),
                #     # Add more fields as needed
                # })
            
        else:
            # BASIC EXTRACTION - Null or Invalid ConfigurationIntent
            if config_intent is None:
                csv_row['config_status'] = 'Null ConfigurationIntent'
                print(f"   Using BASIC extraction for {serial_number} (null ConfigurationIntent)")
            else:
                csv_row['config_status'] = 'Invalid ConfigurationIntent'
                print(f"   Using BASIC extraction for {serial_number} (invalid ConfigurationIntent)")
            
            # CUSTOMIZE THIS SECTION - Add basic fields that exist regardless of ConfigurationIntent
            # These are fields at the root level of the data object
            # Example:
            # csv_row.update({
            #     'device_id': data.get('device_id', ''),
            #     'timestamp': data.get('timestamp', ''),
            #     'status': data.get('status', ''),
            #     # Add more basic fields as needed
            # })
        
        # =================================================================
        # COMMON FIELDS - Always extracted regardless of ConfigurationIntent status
        # =================================================================
        
        # Add fields that should always be extracted for both valid and null devices
        # These would be fields that exist at the root level for all devices
        # Example:
        # csv_row['api_version'] = data.get('api_version', '')
        # csv_row['device_serial'] = data.get('device_serial', serial_number)
        
        return csv_row
        
    except Exception as e:
        print(f"⚠️  CSV extraction failed for {serial_number}: {e}")
        print(f"   Data type: {type(data)}")
        if data:
            print(f"   Data keys: {list(data.keys())[:10]}")  # Show first 10 keys
        print(f"   Skipping CSV data for this serial number")
        return None

def create_excel_file(project_name, all_csv_data, output_dir):
    """
    Create consolidated Excel file from all collected data with proper hyperlinks.
    
    Args:
        project_name (str): Project name for filename
        all_csv_data (list): List of dictionaries containing row data
        output_dir (Path): Output directory path
    """
    if not all_csv_data:
        print("❌ No data to write to Excel")
        return
    
    print(f"🔧 DEBUG: create_excel_file received {len(all_csv_data)} records")
    for i, record in enumerate(all_csv_data):
        print(f"   Record {i+1}: {record.get('serial_number', 'No serial')}")
    
    # Generate Excel filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"{project_name}_{timestamp}.xlsx"
    excel_path = output_dir / excel_filename
    
    try:
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "API Data"
        
        # Get all unique headers from all rows
        all_headers = set()
        for row in all_csv_data:
            all_headers.update(row.keys())
        
        print(f"🔧 DEBUG: Found {len(all_headers)} unique headers: {sorted(all_headers)}")
        
        # Sort headers for consistent column order
        headers = sorted(all_headers)
        
        # Write headers with formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        print(f"🔧 DEBUG: Writing {len(all_csv_data)} data rows...")
        
        # Write data rows
        for row_idx, data_row in enumerate(all_csv_data, 2):  # Start from row 2
            print(f"   Writing row {row_idx-1}: Serial = {data_row.get('serial_number', 'Unknown')}")
            
            for col_idx, header in enumerate(headers, 1):
                value = data_row.get(header, '')
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Handle different types of hyperlinks
                if isinstance(value, str):
                    # Check for hyperlink with shortname format: "URL|Display Text"
                    if '|' in value and any(value.split('|')[0].startswith(proto) for proto in ['http://', 'https://', 'mailto:', 'tel:', 'ftp://']):
                        url, display_text = value.split('|', 1)  # Split only on first |
                        cell.hyperlink = Hyperlink(ref="", target=url.strip())
                        cell.value = display_text.strip()
                        cell.font = Font(color="0000FF", underline="single")  # Blue underlined
                    # Check if it's a URL that should be a hyperlink (no shortname)
                    elif (value.startswith('http://') or value.startswith('https://') or 
                          value.startswith('mailto:') or value.startswith('tel:') or
                          value.startswith('ftp://')):
                        # Create proper Excel hyperlink object
                        cell.hyperlink = Hyperlink(ref="", target=value)
                        cell.value = value
                        cell.font = Font(color="0000FF", underline="single")  # Blue underlined
                    # Handle Excel formula format (legacy support)
                    elif value.startswith('=HYPERLINK('):
                        cell.value = value
                        cell.font = Font(color="0000FF", underline="single")  # Blue underlined
                    else:
                        cell.value = value
                else:
                    cell.value = value
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(excel_path)
        
        print(f"✅ Excel saved: {excel_path}")
        print(f"📊 Total records: {len(all_csv_data)}")
        print(f"📋 Columns: {len(headers)}")
        print(f"🔧 DEBUG: Excel file should have {len(all_csv_data)} data rows + 1 header row = {len(all_csv_data) + 1} total rows")
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        print("   Falling back to CSV format...")
        create_csv_file(project_name, all_csv_data, output_dir)
def create_csv_file(project_name, all_csv_data, output_dir):
    """
    Create consolidated CSV file from all collected data (fallback option).
    
    Args:
        project_name (str): Project name for filename
        all_csv_data (list): List of dictionaries containing CSV row data
        output_dir (Path): Output directory path
    """
    if not all_csv_data:
        print("❌ No data to write to CSV")
        return
    
    # Generate CSV filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_filename = f"{project_name}_{timestamp}.csv"
    csv_path = output_dir / csv_filename
    
    try:
        # Get all unique headers from all rows
        all_headers = set()
        for row in all_csv_data:
            all_headers.update(row.keys())
        
        # Sort headers for consistent column order
        headers = sorted(all_headers)
        
        # Write CSV file
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            writer.writerows(all_csv_data)
        
        print(f"✅ CSV saved: {csv_path}")
        print(f"📊 Total records: {len(all_csv_data)}")
        print(f"📋 Columns: {len(headers)}")
        
    except Exception as e:
        print(f"❌ Error creating CSV: {e}")

def main():
    """Main execution function."""
    
    # Display current user information
    current_user = getpass.getuser()
    print(f"🔐 Running as user: {current_user}")
    print(f"📄 Output format: Excel (.xlsx)")
    print(f"📦 Dependencies: openpyxl library")
    
    # Get script directory and serial file path
    script_dir = Path(__file__).parent
    serial_file = script_dir / "serials.txt"
    
    # Check if serials.txt exists
    if not serial_file.exists():
        print("❌ Error: 'serials.txt' not found in script directory")
        print(f"📁 Expected location: {serial_file}")
        print("💡 Create a 'serials.txt' file with:")
        print("   Line 1: Project Name")
        print("   Line 2+: Serial numbers (one per line)")
        sys.exit(1)
    
    # Read serial numbers and project name
    project_name, serial_numbers = read_serial_numbers(serial_file)
    
    # Pre-check all serials for valid ConfigurationIntent
    valid_serials, invalid_serials = check_configuration_intent(serial_numbers, BASE_URL_TEMPLATE)
    
    # Create output directory in same folder as script
    safe_project_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    output_dir = script_dir / safe_project_name.replace(' ', '_')
    output_dir.mkdir(exist_ok=True)
    
    print(f"📁 Output directory: {output_dir}")
    print(f"🚀 Starting JSON processing for ALL {len(serial_numbers)} serial numbers...")
    print(f"📊 Will include {len(valid_serials)} valid serials in Excel output\n")
    
    # Process ALL serial numbers (for JSON files)
    all_csv_data = []
    successful_count = 0
    
    for i, serial_number in enumerate(serial_numbers, 1):
        print(f"[{i}/{len(serial_numbers)}] Processing: {serial_number}")
        
        # Always fetch JSON data for every serial
        data = fetch_json_data(serial_number, BASE_URL_TEMPLATE)
        
        if data is not None:
            # Always print pretty JSON to terminal
            print("📄 JSON Response:")
            print("-" * 40)
            print(json.dumps(data, indent=2, ensure_ascii=False))
            print("-" * 40)
            
            # Always create pretty JSON file
            create_pretty_json_file(serial_number, data, project_name, output_dir)
            
            # Extract CSV data for ALL serials (conditional extraction based on ConfigurationIntent status)
            csv_row = extract_csv_data(serial_number, data, valid_serials)
            if csv_row is not None:
                all_csv_data.append(csv_row)
                if serial_number in valid_serials:
                    print(f"✅ Full CSV data extracted for {serial_number}")
                else:
                    print(f"✅ Basic CSV data extracted for {serial_number}")
            else:
                print(f"⚠️  CSV extraction failed for {serial_number} (unexpected error)")
            
            successful_count += 1
        
        print()  # Empty line for readability
    
    # Create consolidated Excel file (or CSV as fallback)
    print(f"\n📊 Processing {len(all_csv_data)} records for Excel creation...")
    for i, record in enumerate(all_csv_data):
        print(f"   Record {i+1}: Serial = {record.get('serial_number', 'Unknown')}")
    
    if all_csv_data:
        if EXCEL_AVAILABLE:
            create_excel_file(project_name, all_csv_data, output_dir)
        else:
            create_csv_file(project_name, all_csv_data, output_dir)
    else:
        print("⚠️  No data collected - file creation skipped")
        print("   This could be due to:")
        print("   • All API requests failed")
        print("   • Data extraction errors (missing/mismatched keys)")
        print("   • Customization needed in extract_csv_data() function")
    
    # Summary
    print("="*50)
    print("📋 SUMMARY")
    print("="*50)
    print(f"Project: {project_name}")
    print(f"Total serial numbers submitted: {len(serial_numbers)}")
    print(f"Valid serials processed: {len(valid_serials)}")
    print(f"Invalid serials skipped: {len(invalid_serials)}")
    print(f"Successful retrievals: {successful_count}")
    print(f"Failed retrievals: {len(valid_serials) - successful_count}")
    print(f"Excel/CSV records created: {len(all_csv_data)}")
    print(f"Output directory: {output_dir}")
    
    if invalid_serials:
        print(f"\n🚫 Skipped serials (null ConfigurationIntent):")
        for serial in invalid_serials:
            print(f"   • {serial}")
    print("="*50)

if __name__ == "__main__":
    main()
