#!/usr/bin/env python3
"""
Generic API JSON Scraper with Excel Export
Version: 2.1
Description: Fetches JSON data from API endpoints using serial numbers,
             creates pretty-printed JSON files, and consolidates data into Excel.
             Processes ALL serials with conditional extraction (full vs basic).
             NO SERIALS ARE SKIPPED - all get processed for CSV/Excel output.
             
             NEW: Self-learning device type discovery and JSON structure analysis.
             Script automatically discovers new device types and updates itself.

Usage: python script.py
Requirements: 'serials.txt' file in same directory as script
             pip install openpyxl

Key Features:
- Pre-checks ConfigurationIntent status for all serials
- Processes ALL serials regardless of ConfigurationIntent status
- Full extraction for valid ConfigurationIntent
- Basic extraction for null/invalid ConfigurationIntent
- Creates JSON files for every serial number
- Consolidated Excel/CSV output with all processed serials
- Configurable debug mode for troubleshooting
- Device type discovery and JSON structure analysis
- Self-modifying script with persistent device type dictionary
"""

import subprocess
import json
import csv
import os
import sys
import getpass
import shutil
from datetime import datetime
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.hyperlink import Hyperlink

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Install with: pip install openpyxl")
    print("   Falling back to CSV format...")

# Script Configuration
SCRIPT_VERSION = "2.1"
BASE_URL_TEMPLATE = "https://acme.com/sn="  # Modify this for your API endpoint

# Debug Configuration - Set to True for verbose debugging
DEBUG_MODE = True  # Change to False to reduce console output

# ================================================================
# SELF-LEARNING DEVICE TYPE DISCOVERY SYSTEM
# This dictionary is automatically updated when new device types are discovered
# DO NOT MANUALLY EDIT - the script will modify this section automatically
# ================================================================
DISCOVERED_DEVICE_TYPES = {
    # Format: 'device_type': {'count': number, 'last_seen': 'YYYY-MM-DD', 'first_seen': 'YYYY-MM-DD'}
    # Example entries (will be auto-populated):
    # 'router': {'count': 15, 'last_seen': '2025-01-20', 'first_seen': '2025-01-15'},
    # 'switch': {'count': 8, 'last_seen': '2025-01-19', 'first_seen': '2025-01-10'},
}


def analyze_json_structure(data, device_type, max_depth=10):
    """
    Analyze JSON structure and document all available keys and access patterns.
    Creates comprehensive documentation for accessing nested data.
    
    Args:
        data (dict): JSON data to analyze
        device_type (str): Device type for categorization
        max_depth (int): Maximum traversal depth (default: 10)
    
    Returns:
        dict: Analysis results with paths, access patterns, and statistics
    """
    if not isinstance(data, dict):
        return {'error': 'Data is not a dictionary', 'data_type': str(type(data))}
    
    analysis = {
        'device_type': device_type,
        'analysis_timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'total_keys': 0,
        'max_depth_found': 0,
        'paths': {},
        'access_patterns': [],
        'data_types': {},
        'sample_values': {}
    }
    
    def traverse_json(obj, path="data", current_depth=0):
        """Recursively traverse JSON and document all paths."""
        if current_depth > max_depth:
            return
        
        analysis['max_depth_found'] = max(analysis['max_depth_found'], current_depth)
        
        if isinstance(obj, dict):
            for key, value in obj.items():
                current_path = f"{path}['{key}']" if path != "data" else f"data['{key}']"
                analysis['total_keys'] += 1
                
                # Document the path and access pattern
                analysis['paths'][current_path] = {
                    'depth': current_depth,
                    'data_type': str(type(value).__name__),
                    'access_code': f"value = {current_path}",
                    'safe_access_code': f"value = {path}.get('{key}', '') if isinstance({path}, dict) else ''"
                }
                
                # Store data type
                analysis['data_types'][current_path] = str(type(value).__name__)
                
                # Store sample value (truncated if too long)
                if isinstance(value, (str, int, float, bool)):
                    sample = str(value)
                    analysis['sample_values'][current_path] = sample[:100] + '...' if len(sample) > 100 else sample
                elif value is None:
                    analysis['sample_values'][current_path] = 'null'
                elif isinstance(value, (list, dict)):
                    analysis['sample_values'][current_path] = f"{type(value).__name__} with {len(value)} items"
                
                # Continue traversing
                if isinstance(value, dict) and current_depth < max_depth:
                    traverse_json(value, current_path, current_depth + 1)
                elif isinstance(value, list) and value and current_depth < max_depth:
                    # Analyze first item in list if it exists
                    list_path = f"{current_path}[0]"
                    analysis['paths'][list_path] = {
                        'depth': current_depth + 1,
                        'data_type': f"list_item_{type(value[0]).__name__}",
                        'access_code': f"value = {current_path}[0] if {current_path} else None",
                        'safe_access_code': f"value = {current_path}[0] if isinstance({current_path}, list) and len({current_path}) > 0 else None"
                    }
                    if isinstance(value[0], dict):
                        traverse_json(value[0], list_path, current_depth + 2)
        
        elif isinstance(obj, list) and obj:
            # Handle top-level lists
            for i, item in enumerate(obj[:3]):  # Only analyze first 3 items
                item_path = f"{path}[{i}]"
                if isinstance(item, dict):
                    traverse_json(item, item_path, current_depth + 1)
    
    # Start analysis
    traverse_json(data)
    
    # Generate access patterns documentation
    analysis['access_patterns'] = [
        "# Common Access Patterns for JSON Data",
        "# Replace 'data' with your actual variable name",
        "",
        "# Basic access:",
        "# value = data['key_name']",
        "",
        "# Safe access (prevents KeyError):",
        "# value = data.get('key_name', 'default_value')",
        "",
        "# Nested access:",
        "# value = data['level1']['level2']['level3']",
        "",
        "# Safe nested access:",
        "# value = data.get('level1', {}).get('level2', {}).get('level3', 'default')",
        "",
        "# List access:",
        "# value = data['list_key'][0] if data['list_key'] else None",
        "",
        "# Safe list access:",
        "# value = data.get('list_key', [])[0] if data.get('list_key') else None",
        ""
    ]
    
    return analysis


def save_device_type_analysis(device_type, analysis_data, output_dir):
    """
    Save device type analysis to JSON file with comprehensive documentation.
    
    Args:
        device_type (str): Device type name
        analysis_data (dict): Analysis results from analyze_json_structure
        output_dir (Path): Analysis output directory
    """
    try:
        analysis_dir = output_dir / "device-type-analysis"
        analysis_dir.mkdir(exist_ok=True)
        
        # Save detailed analysis
        analysis_file = analysis_dir / f"{device_type}_analysis.json"
        with open(analysis_file, 'w', encoding='utf-8') as f:
            json.dump(analysis_data, f, indent=2, ensure_ascii=False)
        
        # Save access patterns as text file
        patterns_file = analysis_dir / f"{device_type}_access_patterns.txt"
        with open(patterns_file, 'w', encoding='utf-8') as f:
            f.write(f"# Access Patterns for Device Type: {device_type}\n")
            f.write(f"# Generated: {analysis_data.get('analysis_timestamp', 'Unknown')}\n")
            f.write(f"# Total Keys Found: {analysis_data.get('total_keys', 0)}\n")
            f.write(f"# Max Depth: {analysis_data.get('max_depth_found', 0)}\n\n")
            
            # Write general patterns
            for pattern in analysis_data.get('access_patterns', []):
                f.write(f"{pattern}\n")
            
            f.write("\n# Specific Paths Found in This Device Type:\n\n")
            
            # Write specific paths grouped by depth
            paths_by_depth = defaultdict(list)
            for path, info in analysis_data.get('paths', {}).items():
                paths_by_depth[info['depth']].append((path, info))
            
            for depth in sorted(paths_by_depth.keys()):
                f.write(f"## Depth {depth} Fields:\n")
                for path, info in sorted(paths_by_depth[depth]):
                    f.write(f"# Path: {path}\n")
                    f.write(f"# Type: {info['data_type']}\n")
                    f.write(f"# Access: {info['access_code']}\n")
                    f.write(f"# Safe:   {info['safe_access_code']}\n")
                    sample = analysis_data.get('sample_values', {}).get(path, 'No sample')
                    f.write(f"# Sample: {sample}\n")
                    f.write("\n")
                f.write("\n")
        
        debug_print(f"Saved analysis for device type '{device_type}' to {analysis_dir}")
        
    except Exception as e:
        print(f"⚠️  Error saving device type analysis for '{device_type}': {e}")


def classify_device_type(data):
    """
    ================================================================
    DEVICE TYPE CLASSIFICATION - CUSTOMIZE THIS SECTION
    TODO: Replace 'device_type' with your actual classification key
    
    This is a PLACEHOLDER function that needs to be customized based on
    your actual JSON structure. Look for a key-value pair that identifies
    different types of devices in your API responses.
    
    Common classification keys to look for:
    - data.get('device_type')
    - data.get('product_type') 
    - data.get('model')
    - data.get('category')
    - data.get('class')
    - data.get('hardware_type')
    
    You may need to use nested keys like:
    - data.get('device_info', {}).get('type')
    - data.get('system', {}).get('device_type')
    ================================================================
    
    Args:
        data (dict): JSON data from API response
    
    Returns:
        str: Device type classification
    """
    if not isinstance(data, dict):
        return 'unknown'
    
    # ================================================================
    # PLACEHOLDER CLASSIFICATION LOGIC - CUSTOMIZE THIS!
    # Replace this with your actual device type identification logic
    # ================================================================
    
    # Example placeholder - customize based on your JSON structure:
    device_type = data.get('device_type', 'unknown')
    
    # You might need something like:
    # device_type = data.get('product_info', {}).get('type', 'unknown')
    # OR
    # device_type = data.get('system', {}).get('hardware_type', 'unknown')
    # OR
    # model = data.get('model', '')
    # if 'router' in model.lower():
    #     device_type = 'router'
    # elif 'switch' in model.lower():
    #     device_type = 'switch'
    # else:
    #     device_type = 'unknown'
    
    # Normalize device type (lowercase, no spaces)
    if isinstance(device_type, str):
        return device_type.lower().replace(' ', '_').replace('-', '_')
    
    return 'unknown'


def update_device_type_dictionary(device_type):
    """
    Update the persistent device type dictionary and modify the script file if needed.
    Creates backup before modification.
    
    Args:
        device_type (str): Device type to add/update
    
    Returns:
        bool: True if script was modified, False if no changes needed
    """
    global DISCOVERED_DEVICE_TYPES
    
    current_date = datetime.now().strftime("%Y-%m-%d")
    
    # Check if this is a new device type
    if device_type not in DISCOVERED_DEVICE_TYPES:
        print(f"🔍 NEW DEVICE TYPE DISCOVERED: '{device_type}'")
        
        # Show current state
        print(f"📊 Current Device Types: {dict(DISCOVERED_DEVICE_TYPES)}")
        
        # Add new device type
        DISCOVERED_DEVICE_TYPES[device_type] = {
            'count': 1,
            'first_seen': current_date,
            'last_seen': current_date
        }
        
        print(f"📦 Updated Device Types: {dict(DISCOVERED_DEVICE_TYPES)}")
        
        # Backup and modify script
        backup_script()
        modify_script_with_new_device_type()
        
        return True
    else:
        # Update existing device type
        DISCOVERED_DEVICE_TYPES[device_type]['count'] += 1
        DISCOVERED_DEVICE_TYPES[device_type]['last_seen'] = current_date
        
        return False


def backup_script():
    """
    Create a backup of the current script before modification.
    Adds header comments explaining the backup reason.
    """
    try:
        script_path = Path(__file__)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = script_path.parent / f"{script_path.stem}_backup_{timestamp}{script_path.suffix}"
        
        # Read current script
        with open(script_path, 'r', encoding='utf-8') as f:
            current_content = f.read()
        
        # Create backup header
        backup_header = f'''#!/usr/bin/env python3
"""
BACKUP REASON: New device type discovered
BACKUP TIMESTAMP: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
PREVIOUS DEVICE TYPES: {dict(DISCOVERED_DEVICE_TYPES)}
BACKUP CREATED BY: Automatic device type discovery system

This is an automatic backup created before script self-modification.
The original script discovered a new device type and updated itself.
"""

'''
        
        # Write backup with header
        with open(backup_path, 'w', encoding='utf-8') as f:
            f.write(backup_header)
            f.write(current_content)
        
        print(f"💾 Script backed up to: {backup_path}")
        
    except Exception as e:
        print(f"⚠️  Error creating backup: {e}")


def modify_script_with_new_device_type():
    """
    Modify the current script file to update the DISCOVERED_DEVICE_TYPES dictionary.
    """
    try:
        script_path = Path(__file__)
        
        # Read current script
        with open(script_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # Find and replace the DISCOVERED_DEVICE_TYPES section
        new_lines = []
        in_device_types_section = False
        section_found = False
        
        for line in lines:
            if 'DISCOVERED_DEVICE_TYPES = {' in line:
                in_device_types_section = True
                section_found = True
                new_lines.append(line)
                
                # Write updated dictionary
                new_lines.append('    # Format: \'device_type\': {\'count\': number, \'last_seen\': \'YYYY-MM-DD\', \'first_seen\': \'YYYY-MM-DD\'}\n')
                new_lines.append('    # Auto-updated by device discovery system:\n')
                
                for device_type, info in DISCOVERED_DEVICE_TYPES.items():
                    new_lines.append(f"    '{device_type}': {{'count': {info['count']}, 'last_seen': '{info['last_seen']}', 'first_seen': '{info['first_seen']}'}},\n")
                
                continue
            
            elif in_device_types_section and line.strip() == '}':
                in_device_types_section = False
                new_lines.append(line)
                continue
            
            elif not in_device_types_section:
                new_lines.append(line)
        
        if not section_found:
            print("⚠️  Could not find DISCOVERED_DEVICE_TYPES section in script")
            return
        
        # Write updated script
        with open(script_path, 'w', encoding='utf-8') as f:
            f.writelines(new_lines)
        
        print(f"✅ Script updated with new device type dictionary")
        
    except Exception as e:
        print(f"⚠️  Error modifying script: {e}")
    """
    Print debug messages only when DEBUG_MODE is enabled.
    
    Args:
        message (str): Debug message to print
        force (bool): Force print even if DEBUG_MODE is False
    """
    if DEBUG_MODE or force:
        print(f"🔧 DEBUG: {message}")


def read_serial_numbers(file_path):
    """
    Read serial numbers from text file.
    First line is project name, remaining lines are serial numbers.

    Args:
        file_path (str): Path to text file containing serial numbers

    Returns:
        tuple: (project_name, list_of_serial_numbers)
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = [line.strip() for line in f.readlines() if line.strip()]

        if not lines:
            raise ValueError("File is empty")

        project_name = lines[0]
        serial_numbers = lines[1:]

        print(f"📋 Project: {project_name}")
        print(f"📋 Found {len(serial_numbers)} serial numbers to process")
        debug_print(f"Serial numbers: {serial_numbers}")

        return project_name, serial_numbers

    except FileNotFoundError:
        print(f"❌ Error: File '{file_path}' not found")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        sys.exit(1)


def process_all_serials(serial_numbers, base_url, project_name, output_dir):
    """
    Process all serial numbers in a single pass - fetch JSON, categorize ConfigurationIntent,
    create JSON files, extract CSV data, and perform device type analysis all at once. 
    No redundant API calls!
    
    Args:
        serial_numbers (list): List of serial numbers to process
        base_url (str): Base URL template for API
        project_name (str): Project name for JSON file headers
        output_dir (Path): Output directory for JSON files

    Returns:
        tuple: (all_csv_data, full_extraction_serials, basic_extraction_serials, 
                successful_api_calls, failed_api_calls, device_type_stats)
    """
    print("🚀 Processing all serials in single pass (no redundant API calls)...")
    print("=" * 70)

    full_extraction_serials = []    # Valid ConfigurationIntent - full extraction
    basic_extraction_serials = []   # Null/invalid ConfigurationIntent - basic extraction
    all_csv_data = []               # Collected CSV data for all serials
    successful_api_calls = 0
    failed_api_calls = 0
    
    # Device type analysis tracking
    device_type_stats = defaultdict(int)
    device_type_analyses = {}  # Store one analysis per device type

    for i, serial_number in enumerate(serial_numbers, 1):
        print(f"[{i}/{len(serial_numbers)}] Processing: {serial_number}")

        # Fetch JSON data from API
        data = fetch_json_data(serial_number, base_url)

        if data is not None:
            successful_api_calls += 1
            
            # ================================================================
            # DEVICE TYPE DISCOVERY AND ANALYSIS
            # ================================================================
            device_type = classify_device_type(data)
            device_type_stats[device_type] += 1
            
            print(f"   📋 Device Type: '{device_type}'")
            
            # Update persistent device type dictionary (may modify script)
            script_modified = update_device_type_dictionary(device_type)
            if script_modified:
                print(f"   🔄 Script self-modified to include new device type")
            
            # Perform JSON structure analysis (one per device type)
            if device_type not in device_type_analyses:
                print(f"   🔍 Analyzing JSON structure for device type '{device_type}'...")
                analysis = analyze_json_structure(data, device_type)
                device_type_analyses[device_type] = analysis
                save_device_type_analysis(device_type, analysis, output_dir)
                print(f"   📄 Analysis saved for device type '{device_type}'")
            
            # Check ConfigurationIntent status and categorize
            config_intent = data.get('ConfigurationIntent')
            if config_intent is None:
                print("   ➡️ ConfigurationIntent is null - using basic extraction")
                basic_extraction_serials.append(serial_number)
            elif not isinstance(config_intent, dict):
                print(f"   ➡️ ConfigurationIntent is {type(config_intent).__name__} - using basic extraction")
                basic_extraction_serials.append(serial_number)
            else:
                print("   ✅ Valid ConfigurationIntent - using full extraction")
                full_extraction_serials.append(serial_number)

            # Display JSON response (if debug mode)
            if DEBUG_MODE:
                print("   📄 JSON Response:")
                print("   " + "-" * 40)
                for line in json.dumps(data, indent=2, ensure_ascii=False).split('\n'):
                    print(f"   {line}")
                print("   " + "-" * 40)

            # Create pretty JSON file
            create_pretty_json_file(serial_number, data, project_name, output_dir)

        else:
            failed_api_calls += 1
            print(f"   ❌ API call failed - using basic extraction with minimal data")
            basic_extraction_serials.append(serial_number)
            # Default device type for failed API calls
            device_type = 'api_failure'
            device_type_stats[device_type] += 1

        # Extract CSV data (works whether data is None or valid)
        csv_row = extract_csv_data(serial_number, data, full_extraction_serials)
        
        # Add device type to CSV data
        if csv_row is not None:
            csv_row['device_type'] = device_type if 'device_type' in locals() else 'unknown'
            all_csv_data.append(csv_row)
            extraction_type = csv_row.get('extraction_type', 'unknown')
            config_status = csv_row.get('config_status', 'unknown')
            print(f"   ✅ CSV data extracted ({extraction_type} extraction, status: {config_status})")
        else:
            # Emergency fallback (should never happen with fixed logic)
            print(f"   🚨 CRITICAL ERROR: CSV extraction returned None")
            emergency_row = {
                'serial_number': serial_number,
                'device_type': device_type if 'device_type' in locals() else 'unknown',
                'config_status': 'CRITICAL_ERROR',
                'extraction_type': 'emergency_fallback',
                'data_available': 'false',
                'error_message': 'extract_csv_data returned None'
            }
            all_csv_data.append(emergency_row)
            print(f"   🚨 Added emergency fallback row")

        print()  # Empty line for readability

    # Summary after processing
    print("=" * 70)
    print("📊 PROCESSING SUMMARY:")
    print(f"✅ Full extraction serials: {len(full_extraction_serials)}")
    print(f"➡️ Basic extraction serials: {len(basic_extraction_serials)}")
    print(f"🎯 Total serials processed: {len(serial_numbers)}")
    print(f"✅ Successful API calls: {successful_api_calls}")
    print(f"❌ Failed API calls: {failed_api_calls}")
    print(f"📊 CSV records created: {len(all_csv_data)}")
    
    # Device type summary
    print(f"\n🔍 DEVICE TYPE DISCOVERY:")
    for device_type, count in sorted(device_type_stats.items()):
        print(f"   📋 {device_type}: {count} devices")

    if basic_extraction_serials:
        print(f"\n➡️ Serials with basic extraction (null/invalid ConfigurationIntent):")
        for serial in basic_extraction_serials:
            print(f"   • {serial}")

    if full_extraction_serials:
        print(f"\n✅ Serials with full extraction (valid ConfigurationIntent):")
        for serial in full_extraction_serials:
            print(f"   • {serial}")

    print("=" * 70)

    return all_csv_data, full_extraction_serials, basic_extraction_serials, successful_api_calls, failed_api_calls, device_type_stats Basic extraction serials: {len(basic_extraction_serials)}")
    print(f"🎯 Total serials processed: {len(serial_numbers)}")
    print(f"✅ Successful API calls: {successful_api_calls}")
    print(f"❌ Failed API calls: {failed_api_calls}")
    print(f"📊 CSV records created: {len(all_csv_data)}")

    if basic_extraction_serials:
        print(f"\n➡️ Serials with basic extraction (null/invalid ConfigurationIntent):")
        for serial in basic_extraction_serials:
            print(f"   • {serial}")

    if full_extraction_serials:
        print(f"\n✅ Serials with full extraction (valid ConfigurationIntent):")
        for serial in full_extraction_serials:
            print(f"   • {serial}")

    print("=" * 70)

    return all_csv_data, full_extraction_serials, basic_extraction_serials, successful_api_calls, failed_api_calls


def fetch_json_data(serial_number, base_url):
    """
    Fetch JSON data from API using curl with system authentication.
    Enhanced error handling and debugging for troubleshooting.

    Args:
        serial_number (str): Serial number to query
        base_url (str): Base URL template for API

    Returns:
        dict or None: JSON data if successful, None if failed
    """
    url = f"{base_url}{serial_number}"

    try:
        debug_print(f"Fetching data for SN: {serial_number} from {url}")

        # Execute curl command with system authentication
        curl_cmd = [
            'curl',
            '-s',  # Silent mode
            '-L',  # Follow redirects
            '--max-time', '30',  # 30 second timeout
            '--fail',  # Fail on HTTP errors
            '--negotiate',  # Enable SPNEGO/Negotiate authentication (Kerberos/NTLM)
            '--user', ':',  # Use current user credentials (empty user:pass triggers system auth)
            '--ntlm',  # Enable NTLM authentication for Windows domains
            '--anyauth',  # Let curl pick the best auth method available
            '--insecure',  # Skip SSL certificate verification (for testing)
            '--ssl-no-revoke',  # Don't check SSL certificate revocation (Windows)
            '--tlsv1.2',  # Force TLS 1.2 or higher
            url
        ]

        debug_print(f"Executing curl command: {' '.join(curl_cmd[:8])}... [auth params hidden]")

        result = subprocess.run(curl_cmd, capture_output=True, text=True, check=True, encoding='utf-8')

        debug_print(f"API response length: {len(result.stdout)} characters")

        # Parse JSON response
        data = json.loads(result.stdout)
        debug_print(f"Successfully parsed JSON for {serial_number}")
        return data

    except subprocess.CalledProcessError as e:
        print(f"❌ API request failed for {serial_number}: HTTP error (Status: {e.returncode})")
        if e.stderr:
            debug_print(f"STDERR: {e.stderr}")
        debug_print("This might be an authentication issue. Check your domain credentials.")
        return None
    except json.JSONDecodeError as e:
        print(f"❌ Invalid JSON response for {serial_number}: {e}")
        debug_print(f"Raw response: {result.stdout[:200]}...")
        return None
    except Exception as e:
        print(f"❌ Error fetching {serial_number}: {e}")
        debug_print(f"Exception type: {type(e).__name__}")
        return None


def create_pretty_json_file(serial_number, data, project_name, output_dir):
    """
    Create a pretty-printed JSON file with header information.
    Enhanced error handling and debug information.

    Args:
        serial_number (str): Serial number
        data (dict): JSON data from API
        project_name (str): Project name for header
        output_dir (Path): Output directory path
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Create header information
    header_info = {
        "metadata": {
            "project_name": project_name,
            "serial_number": serial_number,
            "timestamp": timestamp,
            "script_version": SCRIPT_VERSION,
            "data_source": f"{BASE_URL_TEMPLATE}{serial_number}"
        },
        "api_data": data
    }

    # Save to JSON file
    file_path = output_dir / f"{serial_number}.json"

    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(header_info, f, indent=2, ensure_ascii=False)
        print(f"✅ Saved JSON: {file_path}")
        debug_print(f"JSON file size: {file_path.stat().st_size} bytes")

    except Exception as e:
        print(f"❌ Error saving JSON for {serial_number}: {e}")
        debug_print(f"Output directory exists: {output_dir.exists()}")
        debug_print(f"Output directory writable: {os.access(output_dir, os.W_OK)}")


def extract_csv_data(serial_number, data, full_extraction_serials):
    """
    Extract specific data fields for CSV export.
    Uses conditional logic: full extraction for valid ConfigurationIntent, basic extraction for null/invalid.
    CRITICAL: This function should NEVER return None - always returns a dictionary.

    Args:
        serial_number (str): Serial number
        data (dict): JSON data from API
        full_extraction_serials (list): List of serials with valid ConfigurationIntent

    Returns:
        dict: Flattened data for CSV row (NEVER returns None)
    """

    debug_print(f"Starting CSV extraction for {serial_number}")

    # Initialize CSV row with serial number (guaranteed field)
    csv_row = {
        'serial_number': serial_number
    }

    try:
        # Check if data is None or empty
        if data is None:
            print(f"⚠️  No data received for {serial_number} - using minimal extraction")
            csv_row.update({
                'config_status': 'API_FAILURE',
                'extraction_type': 'minimal',
                'data_available': 'false'
            })
            debug_print(f"Returning minimal data for {serial_number} due to API failure")
            return csv_row

        if not isinstance(data, dict):
            print(f"⚠️  Invalid data type for {serial_number} - expected dict, got {type(data)}")
            csv_row.update({
                'config_status': 'INVALID_DATA_TYPE',
                'extraction_type': 'minimal',
                'data_available': 'false',
                'data_type_received': str(type(data).__name__)
            })
            debug_print(f"Returning minimal data for {serial_number} due to invalid data type")
            return csv_row

        # =================================================================
        # EXTRACT ALL ROOT-LEVEL DATA (Always done for every serial)
        # =================================================================

        # Get ConfigurationIntent safely
        config_intent = data.get('ConfigurationIntent')
        debug_print(f"ConfigurationIntent type for {serial_number}: {type(config_intent)}")

        # Set config status for tracking
        if config_intent is None:
            csv_row['config_status'] = 'Null_ConfigurationIntent'
        elif not isinstance(config_intent, dict):
            csv_row['config_status'] = 'Invalid_ConfigurationIntent'
        else:
            csv_row['config_status'] = 'Valid_ConfigurationIntent'

        # CUSTOMIZE THIS SECTION - Extract ALL root-level fields for every device
        # These fields should exist at the root level regardless of ConfigurationIntent status
        # Example:
        # csv_row.update({
        #     'device_id': data.get('device_id', ''),
        #     'timestamp': data.get('timestamp', ''),
        #     'status': data.get('status', ''),
        #     'device_type': data.get('device_type', ''),
        #     'last_seen': data.get('last_seen', ''),
        #     'api_version': data.get('api_version', ''),
        #     'device_serial': data.get('device_serial', serial_number),
        #     'response_timestamp': data.get('timestamp', ''),
        #     'device_status': data.get('status', 'unknown'),
        #     # Add ALL your root-level fields here
        # })

        # Placeholder fields for demonstration - replace with your actual fields
        csv_row.update({
            'root_keys_available': ', '.join(list(data.keys())[:5]) if isinstance(data, dict) else '',
            'total_root_keys': len(data) if isinstance(data, dict) else 0,
            'data_available': 'true',
            'processing_timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

        # =================================================================
        # EXTRACT CONFIGURATIONINTENT DATA (Only if valid)
        # =================================================================

        if serial_number in full_extraction_serials and isinstance(config_intent, dict):
            # ADDITIONAL EXTRACTION - Valid ConfigurationIntent data
            csv_row['extraction_type'] = 'full'
            print(f"   Extracting root data + ConfigurationIntent for {serial_number}")
            debug_print(f"ConfigurationIntent keys: {list(config_intent.keys())}")

            # CUSTOMIZE THIS SECTION - Add ConfigurationIntent specific fields
            intent_data = config_intent.get('IntentData', {})
            debug_print(f"IntentData available: {intent_data is not None}")

            # Example ConfigurationIntent extractions:
            # csv_row.update({
            #     'intent_field_1': intent_data.get('your_intent_field_1', ''),
            #     'intent_field_2': intent_data.get('your_intent_field_2', ''),
            #     'nested_intent_field': intent_data.get('nested_object', {}).get('nested_field', ''),
            #     'config_version': config_intent.get('version', ''),
            #     'config_type': config_intent.get('type', ''),
            #     # Add more ConfigurationIntent fields as needed
            # })

            # Placeholder fields for demonstration
            csv_row.update({
                'intent_data_available': 'true' if intent_data else 'false',
                'intent_data_keys_count': len(intent_data) if isinstance(intent_data, dict) else 0
            })

        else:
            # NO ADDITIONAL EXTRACTION - Only root-level data
            csv_row['extraction_type'] = 'basic'
            print(f"   Extracting root data only for {serial_number} (ConfigurationIntent unavailable)")
            
            # No additional fields needed - all root-level data already extracted above

        debug_print(f"Successfully extracted CSV data for {serial_number}: {len(csv_row)} fields")
        return csv_row

    except Exception as e:
        print(f"⚠️  CSV extraction error for {serial_number}: {e}")
        debug_print(f"Exception details: {type(e).__name__}: {str(e)}")
        
        # CRITICAL: Even on error, return a valid dictionary with error information
        csv_row.update({
            'config_status': 'EXTRACTION_ERROR',
            'extraction_type': 'error',
            'data_available': 'error',
            'error_message': str(e)[:100],  # Truncate long error messages
            'processing_timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        
        debug_print(f"Returning error data for {serial_number} - CSV processing will continue")
        return csv_row


def create_excel_file(project_name, all_csv_data, output_dir):
    """
    Create consolidated Excel file from all collected data with proper hyperlinks.
    Enhanced debugging and error handling.

    Args:
        project_name (str): Project name for filename
        all_csv_data (list): List of dictionaries containing row data
        output_dir (Path): Output directory path
    """
    if not all_csv_data:
        print("❌ No data to write to Excel")
        return

    debug_print(f"create_excel_file received {len(all_csv_data)} records")
    
    if DEBUG_MODE:
        for i, record in enumerate(all_csv_data):
            print(f"   Record {i + 1}: {record.get('serial_number', 'No serial')} - Status: {record.get('config_status', 'Unknown')}")

    # Generate Excel filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"{project_name}_{timestamp}.xlsx"
    excel_path = output_dir / excel_filename

    try:
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "API Data"

        # Get all unique headers from all rows
        all_headers = set()
        for row in all_csv_data:
            all_headers.update(row.keys())

        debug_print(f"Found {len(all_headers)} unique headers: {sorted(all_headers)}")

        # Sort headers for consistent column order (serial_number first)
        headers = ['serial_number'] + sorted([h for h in all_headers if h != 'serial_number'])

        # Write headers with formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        debug_print(f"Writing {len(all_csv_data)} data rows...")

        # Write data rows
        for row_idx, data_row in enumerate(all_csv_data, 2):  # Start from row 2
            debug_print(f"Writing row {row_idx - 1}: Serial = {data_row.get('serial_number', 'Unknown')}")

            for col_idx, header in enumerate(headers, 1):
                value = data_row.get(header, '')
                cell = worksheet.cell(row=row_idx, column=col_idx)

                # Handle different types of hyperlinks
                if isinstance(value, str):
                    # Check for hyperlink with shortname format: "URL|Display Text"
                    if '|' in value and any(value.split('|')[0].startswith(proto) for proto in
                                            ['http://', 'https://', 'mailto:', 'tel:', 'ftp://']):
                        url, display_text = value.split('|', 1)  # Split only on first |
                        cell.hyperlink = Hyperlink(ref="", target=url.strip())
                        cell.value = display_text.strip()
                        cell.font = Font(color="0000FF", underline="single")  # Blue underlined
                    # Check if it's a URL that should be a hyperlink (no shortname)
                    elif (value.startswith('http://') or value.startswith('https://') or
                          value.startswith('mailto:') or value.startswith('tel:') or
                          value.startswith('ftp://')):
                        # Create proper Excel hyperlink object
                        cell.hyperlink = Hyperlink(ref="", target=value)
                        cell.value = value
                        cell.font = Font(color="0000FF", underline="single")  # Blue underlined
                    # Handle Excel formula format (legacy support)
                    elif value.startswith('=HYPERLINK('):
                        cell.value = value
                        cell.font = Font(color="0000FF", underline="single")  # Blue underlined
                    else:
                        cell.value = value
                else:
                    cell.value = value

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        workbook.save(excel_path)

        print(f"✅ Excel saved: {excel_path}")
        print(f"📊 Total records: {len(all_csv_data)}")
        print(f"📋 Columns: {len(headers)}")
        debug_print(f"Excel file should have {len(all_csv_data)} data rows + 1 header row = {len(all_csv_data) + 1} total rows")

    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        debug_print(f"Excel error details: {type(e).__name__}: {str(e)}")
        print("   Falling back to CSV format...")
        create_csv_file(project_name, all_csv_data, output_dir)


def create_csv_file(project_name, all_csv_data, output_dir):
    """
    Create consolidated CSV file from all collected data (fallback option).
    Enhanced error handling and debugging.

    Args:
        project_name (str): Project name for filename
        all_csv_data (list): List of dictionaries containing CSV row data
        output_dir (Path): Output directory path
    """
    if not all_csv_data:
        print("❌ No data to write to CSV")
        return

    # Generate CSV filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_filename = f"{project_name}_{timestamp}.csv"
    csv_path = output_dir / csv_filename

    try:
        # Get all unique headers from all rows
        all_headers = set()
        for row in all_csv_data:
            all_headers.update(row.keys())

        # Sort headers for consistent column order (serial_number first)
        headers = ['serial_number'] + sorted([h for h in all_headers if h != 'serial_number'])
        debug_print(f"CSV headers: {headers}")

        # Write CSV file
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            writer.writerows(all_csv_data)

        print(f"✅ CSV saved: {csv_path}")
        print(f"📊 Total records: {len(all_csv_data)}")
        print(f"📋 Columns: {len(headers)}")

    except Exception as e:
        print(f"❌ Error creating CSV: {e}")
        debug_print(f"CSV error details: {type(e).__name__}: {str(e)}")


def main():
    """
    Main execution function.
    Enhanced with comprehensive error handling and debug information.
    """

    # Display current user information
    current_user = getpass.getuser()
    print(f"🔐 Running as user: {current_user}")
    print(f"📄 Output format: Excel (.xlsx)")
    print(f"📦 Dependencies: openpyxl library")
    print(f"🔧 Debug mode: {'ENABLED' if DEBUG_MODE else 'DISABLED'}")
    print(f"📝 Script version: {SCRIPT_VERSION}")

    # Get script directory and serial file path
    script_dir = Path(__file__).parent
    serial_file = script_dir / "serials.txt"
    debug_print(f"Script directory: {script_dir}")
    debug_print(f"Serial file path: {serial_file}")

    # Check if serials.txt exists
    if not serial_file.exists():
        print("❌ Error: 'serials.txt' not found in script directory")
        print(f"📁 Expected location: {serial_file}")
        print("💡 Create a 'serials.txt' file with:")
        print("   Line 1: Project Name")
        print("   Line 2+: Serial numbers (one per line)")
        sys.exit(1)

    # Read serial numbers and project name
    project_name, serial_numbers = read_serial_numbers(serial_file)

    # Create output directory in same folder as script
    safe_project_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    output_dir = script_dir / safe_project_name.replace(' ', '_')
    output_dir.mkdir(exist_ok=True)

    print(f"📁 Output directory: {output_dir}")
    print(f"🚀 Starting single-pass processing for ALL {len(serial_numbers)} serial numbers...")
    print(f"🎯 EFFICIENT: One API call per serial (no redundant requests)\n")

    # Process ALL serial numbers in a single pass (NO REDUNDANT API CALLS)
    all_csv_data, full_extraction_serials, basic_extraction_serials, successful_api_calls, failed_api_calls, device_type_stats = process_all_serials(
        serial_numbers, BASE_URL_TEMPLATE, project_name, output_dir
    )

    # Create consolidated Excel file (or CSV as fallback)
    print(f"\n📊 Creating consolidated output file...")
    print(f"🎯 Processing {len(all_csv_data)} records for Excel creation...")
    
    debug_print("Final CSV data summary:")
    for i, record in enumerate(all_csv_data):
        debug_print(f"   Record {i + 1}: Serial = {record.get('serial_number', 'Unknown')}, Status = {record.get('config_status', 'Unknown')}")

    if all_csv_data:
        if EXCEL_AVAILABLE:
            create_excel_file(project_name, all_csv_data, output_dir)
        else:
            create_csv_file(project_name, all_csv_data, output_dir)
    else:
        print("❌ CRITICAL ERROR: No CSV data collected!")
        print("   This should never happen with the fixed logic.")
        print("   Check the extract_csv_data() function for issues.")

    # Enhanced Summary with Device Type Information
    print("=" * 70)
    print("📋 FINAL SUMMARY")
    print("=" * 70)
    print(f"Project: {project_name}")
    print(f"Total serial numbers submitted: {len(serial_numbers)}")
    print(f"🎯 Full extraction serials: {len(full_extraction_serials)}")
    print(f"🎯 Basic extraction serials: {len(basic_extraction_serials)}")
    print(f"✅ Successful API calls: {successful_api_calls}")
    print(f"❌ Failed API calls: {failed_api_calls}")
    print(f"📊 CSV/Excel records created: {len(all_csv_data)}")
    print(f"📁 Output directory: {output_dir}")

    # Verify no serials were skipped
    if len(all_csv_data) == len(serial_numbers):
        print(f"✅ SUCCESS: ALL {len(serial_numbers)} serials processed (no skipping)")
    else:
        print(f"❌ WARNING: Expected {len(serial_numbers)} records, got {len(all_csv_data)}")
        print("   This indicates a logic error in the processing loop.")

    # Breakdown by extraction type
    full_extraction_count = sum(1 for row in all_csv_data if row.get('extraction_type') == 'full')
    basic_extraction_count = sum(1 for row in all_csv_data if row.get('extraction_type') == 'basic')
    error_extraction_count = sum(1 for row in all_csv_data if row.get('extraction_type') in ['error', 'minimal', 'emergency_fallback'])

    print(f"\n📈 EXTRACTION BREAKDOWN:")
    print(f"   Full extraction records: {full_extraction_count}")
    print(f"   Basic extraction records: {basic_extraction_count}")
    print(f"   Error/minimal records: {error_extraction_count}")

    # Status breakdown
    status_counts = {}
    for row in all_csv_data:
        status = row.get('config_status', 'Unknown')
        status_counts[status] = status_counts.get(status, 0) + 1

    print(f"\n📊 STATUS BREAKDOWN:")
    for status, count in sorted(status_counts.items()):
        print(f"   {status}: {count}")

    # Device type breakdown
    print(f"\n🔍 DEVICE TYPE BREAKDOWN:")
    for device_type, count in sorted(device_type_stats.items()):
        print(f"   📋 {device_type}: {count} devices")

    # Analysis files created
    analysis_dir = output_dir / "device-type-analysis"
    if analysis_dir.exists():
        analysis_files = list(analysis_dir.glob("*_analysis.json"))
        pattern_files = list(analysis_dir.glob("*_access_patterns.txt"))
        print(f"\n📄 DEVICE TYPE ANALYSIS FILES CREATED:")
        print(f"   📊 Analysis files: {len(analysis_files)}")
        print(f"   📝 Pattern files: {len(pattern_files)}")
        print(f"   📁 Analysis directory: {analysis_dir}")

    # Self-modification summary
    if any(device_type not in ['unknown', 'api_failure'] for device_type in device_type_stats.keys()):
        print(f"\n🔄 SELF-LEARNING SUMMARY:")
        print(f"   📦 Total device types in dictionary: {len(DISCOVERED_DEVICE_TYPES)}")
        print(f"   🆕 Device types discovered this run: {len([dt for dt in device_type_stats.keys() if dt not in ['unknown', 'api_failure']])}")
        print(f"   💾 Script may have been auto-updated with new device types")

    print("=" * 70)

    if DEBUG_MODE:
        print("\n🔧 DEBUG MODE was enabled during this run.")
        print("   To reduce console output, set DEBUG_MODE = False in the script.")
        print("\n🔍 DEVICE TYPE CLASSIFICATION:")
        print("   📝 Remember to customize the classify_device_type() function")
        print("   📝 Current classification uses: data.get('device_type', 'unknown')")
        print("   📝 Check the device-type-analysis files to find better classification keys")


if __name__ == "__main__":
    main()
