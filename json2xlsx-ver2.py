#!/usr/bin/env python3
"""
Generic API JSON Scraper with Excel Export
Version: 1.0
Description: Fetches JSON data from API endpoints using serial numbers,
             creates pretty-printed JSON files, and consolidates data into Excel.
             
Usage: python script.py
Requirements: 'serials.txt' file in same directory as script
             pip install openpyxl
"""

import subprocess
import json
import csv
import os
import sys
import getpass
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.hyperlink import Hyperlink
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Install with: pip install openpyxl")
    print("   Falling back to CSV format...")

# Script Configuration
SCRIPT_VERSION = "1.0"
BASE_URL_TEMPLATE = "https://acme.com/sn="  # Modify this for your API endpoint

def read_serial_numbers(file_path):
    """
    Read serial numbers from text file.
    First line is project name, remaining lines are serial numbers.
    
    Args:
        file_path (str): Path to text file containing serial numbers
        
    Returns:
        tuple: (project_name, list_of_serial_numbers)
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = [line.strip() for line in f.readlines() if line.strip()]
        
        if not lines:
            raise ValueError("File is empty")
            
        project_name = lines[0]
        serial_numbers = lines[1:]
        
        print(f"Project: {project_name}")
        print(f"Found {len(serial_numbers)} serial numbers")
        
        return project_name, serial_numbers
        
    except FileNotFoundError:
        print(f"❌ Error: File '{file_path}' not found")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        sys.exit(1)

def fetch_json_data(serial_number, base_url):
    """
    Fetch JSON data from API using curl with system authentication.
    
    Args:
        serial_number (str): Serial number to query
        base_url (str): Base URL template for API
        
    Returns:
        dict or None: JSON data if successful, None if failed
    """
    url = f"{base_url}{serial_number}"
    
    try:
        print(f"Fetching data for SN: {serial_number}")
        
        # Execute curl command with system authentication
        curl_cmd = [
            'curl',
            '-s',  # Silent mode
            '-L',  # Follow redirects
            '--max-time', '30',  # 30 second timeout
            '--fail',  # Fail on HTTP errors
            '--negotiate',  # Enable SPNEGO/Negotiate authentication (Kerberos/NTLM)
            '--user', ':',  # Use current user credentials (empty user:pass triggers system auth)
            '--ntlm',  # Enable NTLM authentication for Windows domains
            '--anyauth',  # Let curl pick the best auth method available
            '--insecure',  # Skip SSL certificate verification (for testing)
            '--ssl-no-revoke',  # Don't check SSL certificate revocation (Windows)
            '--tlsv1.2',  # Force TLS 1.2 or higher
            url
        ]
        
        result = subprocess.run(curl_cmd, capture_output=True, text=True, check=True, encoding='utf-8')
        
        # Parse JSON response
        data = json.loads(result.stdout)
        return data
        
    except subprocess.CalledProcessError as e:
        print(f"❌ API request failed for {serial_number}: HTTP error (Status: {e.returncode})")
        print(f"   This might be an authentication issue. Check your domain credentials.")
        return None
    except json.JSONDecodeError as e:
        print(f"❌ Invalid JSON response for {serial_number}: {e}")
        return None
    except Exception as e:
        print(f"❌ Error fetching {serial_number}: {e}")
        return None

def create_pretty_json_file(serial_number, data, project_name, output_dir):
    """
    Create a pretty-printed JSON file with header information.
    
    Args:
        serial_number (str): Serial number
        data (dict): JSON data from API
        project_name (str): Project name for header
        output_dir (Path): Output directory path
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Create header information
    header_info = {
        "metadata": {
            "project_name": project_name,
            "serial_number": serial_number,
            "timestamp": timestamp,
            "script_version": SCRIPT_VERSION,
            "data_source": f"{BASE_URL_TEMPLATE}{serial_number}"
        },
        "api_data": data
    }
    
    # Save to JSON file
    file_path = output_dir / f"{serial_number}.json"
    
    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(header_info, f, indent=2, ensure_ascii=False)
        print(f"✅ Saved JSON: {file_path}")
        
    except Exception as e:
        print(f"❌ Error saving JSON for {serial_number}: {e}")

def extract_csv_data(serial_number, data):
    """
    Extract specific data fields for CSV export.
    
    CUSTOMIZE THIS FUNCTION FOR YOUR API DATA STRUCTURE
    
    Args:
        serial_number (str): Serial number
        data (dict): JSON data from API
        
    Returns:
        dict or None: Flattened data for CSV row, or None if extraction fails
    """
    
    try:
        csv_row = {
            'serial_number': serial_number
        }
        
        # =================================================================
        # CUSTOMIZE THIS SECTION FOR YOUR SPECIFIC API DATA STRUCTURE
        # =================================================================
        
        # Example 1: Direct field extraction with hyperlinks
        # Uncomment and modify these based on your JSON structure
        """
        # Regular fields (no hyperlinks)
        csv_row.update({
            'name': data.get('name', ''),
            'model': data.get('model', ''),
            'location': data.get('location', ''),
            'last_seen': data.get('last_seen', ''),
        })
        
        # Method 1: URLs with shortnames (Format: "URL|Display Text")
        status_value = data.get('status', '')
        if status_value:
            csv_row['status_portal'] = f'https://status-portal.com/device/{serial_number}|Check Status'
        
        ip_value = data.get('ip_address', '')
        if ip_value:
            csv_row['ip_address'] = ip_value  # Regular field
            csv_row['web_interface'] = f'https://{ip_value}|Open Web UI'  # Hyperlink with shortname
        
        email_value = data.get('email', '')
        if email_value:
            csv_row['email'] = f'mailto:{email_value}|Send Email'  # Hyperlink with shortname
        
        phone_value = data.get('phone', '')
        if phone_value:
            csv_row['phone'] = f'tel:{phone_value}|Call Device'  # Hyperlink with shortname
        
        # Method 2: Direct URLs (no shortname - shows full URL)
        model_value = data.get('model', '')
        if model_value:
            csv_row['model_docs'] = f'https://docs.vendor.com/{model_value}'  # Shows full URL
        """
        
        # Example 2: Nested field extraction with hyperlinks
        # For nested JSON like: {"device": {"info": {"model": "ABC123"}}}
        """
        device_info = data.get('device', {}).get('info', {})
        
        # Regular nested fields
        csv_row.update({
            'device_type': device_info.get('type', ''),
            'firmware_version': device_info.get('firmware', ''),
        })
        
        # Nested fields with shortnames
        model_value = device_info.get('model', '')
        if model_value:
            csv_row['device_model'] = model_value  # Display value
            csv_row['model_docs'] = f'https://support.vendor.com/model/{model_value}|View Manual'  # Shortname
        
        # Using nested management IP for web interface link
        mgmt_ip = data.get('management', {}).get('ip_address', '')
        if mgmt_ip:
            csv_row['management_ip'] = mgmt_ip  # Display IP
            csv_row['mgmt_interface'] = f'https://{mgmt_ip}|Manage Device'  # Shortname
        """
        
        # Example 3: Array/list handling with hyperlinks
        # For JSON with arrays: {"tags": ["tag1", "tag2", "tag3"]}
        """
        # Regular array handling
        tags = data.get('tags', [])
        csv_row['tags'] = ', '.join(tags) if tags else ''
        
        # Create search URLs for tags
        if tags:
            # Single search URL for all tags
            tags_string = ','.join(tags)
            csv_row['tags_search'] = f'https://search.company.com/tags/{tags_string}'
            
            # Or individual tag search URLs (pick one approach)
            for i, tag in enumerate(tags[:3]):  # Limit to first 3 tags
                csv_row[f'tag_{i+1}_search'] = f'https://search.company.com/tag/{tag}'
        """
        
        # Example 4: Creating hyperlinks from multiple data points
        # Combine different fields to create useful links
        """
        # Create URLs with descriptive shortnames
        asset_id = data.get('asset_id', '')
        if asset_id:
            csv_row['asset_id'] = asset_id  # Display value
            csv_row['asset_portal'] = f'https://assets.company.com/device/{asset_id}|View Asset Details'  # Shortname
        
        # Create monitoring dashboard link using multiple fields
        device_type = data.get('type', '')
        site_code = data.get('site_code', '')
        if device_type and site_code:
            csv_row['monitoring_dashboard'] = f'https://monitor.com/site/{site_code}/type/{device_type}|Monitor Dashboard'
        
        # Create support ticket link with pre-filled info
        model = data.get('model', '')
        if model:
            csv_row['create_support_ticket'] = f'https://support.com/new?serial={serial_number}&model={model}|Create Ticket'
        """
        
        # =================================================================
        # GENERIC FALLBACK: Extract all top-level fields
        # Remove this section once you customize the specific extractions above
        # =================================================================
        
        # This extracts ALL top-level keys from the JSON
        # Customize this by commenting out and using specific extractions above
        """
        for key, value in data.items():
            if isinstance(value, (str, int, float, bool)):
                csv_row[key] = value
            elif isinstance(value, (list, dict)):
                # Convert complex types to string representation
                csv_row[key] = str(value)
            else:
                csv_row[key] = str(value)
        """
        
        return csv_row
        
    except Exception as e:
        print(f"⚠️  CSV extraction failed for {serial_number}: {e}")
        print(f"   Skipping CSV data for this serial number")
        return None

def create_excel_file(project_name, all_csv_data, output_dir):
    """
    Create consolidated Excel file from all collected data with proper hyperlinks.
    
    Args:
        project_name (str): Project name for filename
        all_csv_data (list): List of dictionaries containing row data
        output_dir (Path): Output directory path
    """
    if not all_csv_data:
        print("❌ No data to write to Excel")
        return
    
    # Generate Excel filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"{project_name}_{timestamp}.xlsx"
    excel_path = output_dir / excel_filename
    
    try:
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "API Data"
        
        # Get all unique headers from all rows
        all_headers = set()
        for row in all_csv_data:
            all_headers.update(row.keys())
        
        # Sort headers for consistent column order
        headers = sorted(all_headers)
        
        # Write headers with formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Write data rows
        for row_idx, data_row in enumerate(all_csv_data, 2):  # Start from row 2
            for col_idx, header in enumerate(headers, 1):
                value = data_row.get(header, '')
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Handle different types of hyperlinks
                if isinstance(value, str):
                    # Check for hyperlink with shortname format: "URL|Display Text"
                    if '|' in value and any(value.split('|')[0].startswith(proto) for proto in ['http://', 'https://', 'mailto:', 'tel:', 'ftp://']):
                        url, display_text = value.split('|', 1)  # Split only on first |
                        cell.hyperlink = Hyperlink(ref="", target=url.strip())
                        cell.value = display_text.strip()
                        cell.font = Font(color="0000FF", underline="single")  # Blue underlined
                    # Check if it's a URL that should be a hyperlink (no shortname)
                    elif (value.startswith('http://') or value.startswith('https://') or 
                          value.startswith('mailto:') or value.startswith('tel:') or
                          value.startswith('ftp://')):
                        # Create proper Excel hyperlink object
                        cell.hyperlink = Hyperlink(ref="", target=value)
                        cell.value = value
                        cell.font = Font(color="0000FF", underline="single")  # Blue underlined
                    # Handle Excel formula format (legacy support)
                    elif value.startswith('=HYPERLINK('):
                        cell.value = value
                        cell.font = Font(color="0000FF", underline="single")  # Blue underlined
                    else:
                        cell.value = value
                else:
                    cell.value = value
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        workbook.save(excel_path)
        
        print(f"✅ Excel saved: {excel_path}")
        print(f"📊 Total records: {len(all_csv_data)}")
        print(f"📋 Columns: {len(headers)}")
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        print("   Falling back to CSV format...")
        create_csv_file(project_name, all_csv_data, output_dir)
def create_csv_file(project_name, all_csv_data, output_dir):
    """
    Create consolidated CSV file from all collected data (fallback option).
    
    Args:
        project_name (str): Project name for filename
        all_csv_data (list): List of dictionaries containing CSV row data
        output_dir (Path): Output directory path
    """
    if not all_csv_data:
        print("❌ No data to write to CSV")
        return
    
    # Generate CSV filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_filename = f"{project_name}_{timestamp}.csv"
    csv_path = output_dir / csv_filename
    
    try:
        # Get all unique headers from all rows
        all_headers = set()
        for row in all_csv_data:
            all_headers.update(row.keys())
        
        # Sort headers for consistent column order
        headers = sorted(all_headers)
        
        # Write CSV file
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            writer.writerows(all_csv_data)
        
        print(f"✅ CSV saved: {csv_path}")
        print(f"📊 Total records: {len(all_csv_data)}")
        print(f"📋 Columns: {len(headers)}")
        
    except Exception as e:
        print(f"❌ Error creating CSV: {e}")

def main():
    """Main execution function."""
    
    # Display current user information
    current_user = getpass.getuser()
    print(f"🔐 Running as user: {current_user}")
    print(f"📄 Output format: Excel (.xlsx)")
    print(f"📦 Dependencies: openpyxl library")
    
    # Get script directory and serial file path
    script_dir = Path(__file__).parent
    serial_file = script_dir / "serials.txt"
    
    # Check if serials.txt exists
    if not serial_file.exists():
        print("❌ Error: 'serials.txt' not found in script directory")
        print(f"📁 Expected location: {serial_file}")
        print("💡 Create a 'serials.txt' file with:")
        print("   Line 1: Project Name")
        print("   Line 2+: Serial numbers (one per line)")
        sys.exit(1)
    
    # Read serial numbers and project name
    project_name, serial_numbers = read_serial_numbers(serial_file)
    
    # Create output directory in same folder as script
    safe_project_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    output_dir = script_dir / safe_project_name.replace(' ', '_')
    output_dir.mkdir(exist_ok=True)
    
    print(f"📁 Output directory: {output_dir}")
    print(f"🚀 Starting data collection for {len(serial_numbers)} serial numbers...\n")
    
    # Collect data for all serial numbers
    all_csv_data = []
    successful_count = 0
    
    for i, serial_number in enumerate(serial_numbers, 1):
        print(f"[{i}/{len(serial_numbers)}] Processing: {serial_number}")
        
        # Fetch JSON data
        data = fetch_json_data(serial_number, BASE_URL_TEMPLATE)
        
        if data is not None:
            # Print pretty JSON to terminal
            print("📄 JSON Response:")
            print("-" * 40)
            print(json.dumps(data, indent=2, ensure_ascii=False))
            print("-" * 40)
            
            # Create pretty JSON file
            create_pretty_json_file(serial_number, data, project_name, output_dir)
            
            # Extract data for CSV (with error handling)
            csv_row = extract_csv_data(serial_number, data)
            if csv_row is not None:
                all_csv_data.append(csv_row)
            else:
                print(f"⚠️  Skipping CSV data for {serial_number} due to extraction errors")
            
            successful_count += 1
        
        print()  # Empty line for readability
    
    # Create consolidated Excel file (or CSV as fallback)
    if all_csv_data:
        if EXCEL_AVAILABLE:
            create_excel_file(project_name, all_csv_data, output_dir)
        else:
            create_csv_file(project_name, all_csv_data, output_dir)
    else:
        print("⚠️  No data collected - file creation skipped")
        print("   This could be due to:")
        print("   • All API requests failed")
        print("   • Data extraction errors (missing/mismatched keys)")
        print("   • Customization needed in extract_csv_data() function")
    
    # Summary
    print("="*50)
    print("📋 SUMMARY")
    print("="*50)
    print(f"Project: {project_name}")
    print(f"Total serial numbers: {len(serial_numbers)}")
    print(f"Successful retrievals: {successful_count}")
    print(f"Failed retrievals: {len(serial_numbers) - successful_count}")
    print(f"Excel/CSV records created: {len(all_csv_data)}")
    print(f"Output directory: {output_dir}")
    print("="*50)

if __name__ == "__main__":
    main()
